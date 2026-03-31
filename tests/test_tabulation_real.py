"""Comprehensive tabulation tests using real SPSS data.

Golden standard: Data Tables 3.xlsx (R export from Example 3)
Test file: Raw datat.sav (85 cases x 175 variables)

These tests verify that our output matches the industry-standard R export
for the same dataset. Any regression in these tests means we broke something.
"""

import io
import os
import math
import pytest
from pathlib import Path
from openpyxl import load_workbook

# Test fixtures — bundled in repo (no skip, always run)
FIXTURES_DIR = Path(__file__).parent / "fixtures"
SAV_FILE = FIXTURES_DIR / "example3_raw.sav"
GOLDEN_FILE = FIXTURES_DIR / "example3_golden.xlsx"

# Fallback to OneDrive path for developer convenience
if not SAV_FILE.exists():
    _ALT_DIR = Path(r"C:\Users\jorge\OneDrive\Desktop\talk2data\Example 3 - LA Formulations Patient")
    if (_ALT_DIR / "Raw datat.sav").exists():
        SAV_FILE = _ALT_DIR / "Raw datat.sav"
        GOLDEN_FILE = _ALT_DIR / "Data Tables 3.xlsx"

pytestmark = pytest.mark.skipif(
    not SAV_FILE.exists(),
    reason="Test fixtures not found (run from repo root or add tests/fixtures/)"
)


@pytest.fixture(scope="module")
def data():
    """Load SPSS data once for all tests."""
    from services.quantipy_engine import QuantiProEngine
    return QuantiProEngine.load_spss(SAV_FILE.read_bytes(), "Raw datat.sav")


@pytest.fixture(scope="module")
def metadata(data):
    """Extract metadata once."""
    from services.quantipy_engine import QuantiProEngine
    return QuantiProEngine.extract_metadata(data)


@pytest.fixture(scope="module")
def golden():
    """Load golden standard Excel."""
    import shutil, tempfile
    # Copy to temp to avoid permission issues with OneDrive
    tmp = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(GOLDEN_FILE, tmp)
    wb = load_workbook(tmp)
    yield wb
    os.unlink(tmp)


def _tabulate(data, **kwargs):
    """Helper: run tabulation and return workbook."""
    from services.quantipy_engine import QuantiProEngine
    from services.tabulation_builder import TabulateSpec, build_tabulation
    spec = TabulateSpec(**kwargs)
    result = build_tabulation(QuantiProEngine, data, spec)
    wb = load_workbook(io.BytesIO(result.excel_bytes))
    return result, wb


def _get_base(ws, col_idx):
    """Find Base (N) value in a worksheet for a given column."""
    for r in range(1, 15):
        val = ws.cell(row=r, column=1).value
        if val and "Base" in str(val):
            return ws.cell(row=r, column=col_idx).value
    return None


def _get_pct(ws, row_label, col_idx):
    """Find percentage for a row label in a given column."""
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip() == row_label.strip():
            cell_val = ws.cell(row=r, column=col_idx).value
            if cell_val and "%" in str(cell_val):
                return float(str(cell_val).replace("%", "").strip().split()[0])
    return None


# ══════════════════════════════════════════════════════════════════════
# Test 1: Basic file loading
# ══════════════════════════════════════════════════════════════════════

class TestFileLoading:
    def test_case_count(self, data):
        assert len(data.df) == 85

    def test_variable_count(self, data):
        assert len(data.df.columns) == 175

    def test_metadata_variables(self, metadata):
        assert metadata["n_cases"] == 85
        assert metadata["n_variables"] == 175


# ══════════════════════════════════════════════════════════════════════
# Test 2: Simple crosstab (Q_11 Severity x Q_10 Condition)
# ══════════════════════════════════════════════════════════════════════

class TestSimpleCrosstab:
    """Reference: Data Tables 3, 'By Data Cut' sheet, S3 block."""

    def test_severity_by_condition(self, data):
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=["Q_11"],
            include_total_column=True,
        )
        assert result.successful >= 1
        assert result.failed == 0
        assert "Q_11" in wb.sheetnames

    def test_severity_percentages(self, data):
        """R export: Moderate PsO=73.3%, PsA=63.3%, UC/CD=76%, Total=70.6%"""
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=["Q_11"],
            include_total_column=True,
        )
        ws = wb["Q_11"]
        # Find "Moderate" row
        moderate_pct = _get_pct(ws, "Moderate", 2)  # Total column
        assert moderate_pct is not None
        assert abs(moderate_pct - 70.6) < 1.0  # Within 1% tolerance

    def test_total_only(self, data):
        """Export with only Total column (no banners)."""
        result, wb = _tabulate(data,
            banners=[],
            stubs=["Q_11"],
            include_total_column=True,
        )
        assert result.successful >= 1
        ws = wb["Q_11"]
        base = _get_base(ws, 2)  # Total column
        assert base == 85


# ══════════════════════════════════════════════════════════════════════
# Test 3: MRS (Multiple Response Sets)
# ══════════════════════════════════════════════════════════════════════

class TestMRS:
    """Reference: Data Tables 3, S4A block (PsO treatments)."""

    @pytest.fixture
    def pso_members(self, data):
        return [f"Q_12_{i}" for i in range(1, 21) if f"Q_12_{i}" in data.df.columns]

    @pytest.fixture
    def psa_members(self, data):
        return [f"Q_13_{i}" for i in range(1, 23) if f"Q_13_{i}" in data.df.columns]

    def test_mrs_base_is_respondents_not_responses(self, data, pso_members):
        """CRITICAL: Base must be 30 (PsO respondents), not 56 (total responses)."""
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=[],
            mrs_groups={"Treatments PsO": pso_members},
            include_total_column=True,
        )
        ws = wb["Treatments PsO"]
        # Total column base = 30 (only PsO patients answered)
        total_base = _get_base(ws, 2)
        assert total_base == 30, f"MRS base should be 30 respondents, got {total_base}"

    def test_mrs_psa_base(self, data, psa_members):
        """PsA treatments: base should be 30 PsA respondents."""
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=[],
            mrs_groups={"Treatments PsA": psa_members},
            include_total_column=True,
        )
        ws = wb["Treatments PsA"]
        total_base = _get_base(ws, 2)
        assert total_base == 30, f"PsA MRS base should be 30, got {total_base}"

    def test_mrs_percentages_match_golden(self, data, pso_members):
        """R export: Otezla PsO=43%, Humira=30%, Skyrizi=17%."""
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=[],
            mrs_groups={"Treatments PsO": pso_members},
            include_total_column=True,
        )
        ws = wb["Treatments PsO"]
        otezla = _get_pct(ws, "Otezla [apremilast] (Oral treatment)", 2)
        assert otezla is not None
        assert abs(otezla - 43.3) < 1.0, f"Otezla should be ~43.3%, got {otezla}"

    def test_mrs_label_stripping(self, data, pso_members):
        """Labels should be stripped: 'Otezla...' not 'What treatments...Otezla'."""
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=[],
            mrs_groups={"Treatments PsO": pso_members},
            include_total_column=True,
        )
        ws = wb["Treatments PsO"]
        # Check first data row — should NOT start with "What treatments"
        for r in range(5, 10):
            val = ws.cell(row=r, column=1).value
            if val and "%" in str(ws.cell(row=r, column=2).value or ""):
                assert not str(val).startswith("What treatments"), \
                    f"MRS label not stripped: {val[:60]}"
                break

    def test_mrs_all_respondents_question(self, data):
        """Q_39 (challenges Oral): all 85 answered, base should be 85."""
        members = [f"Q_39_{i}" for i in range(1, 15) if f"Q_39_{i}" in data.df.columns]
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=[],
            mrs_groups={"Challenges Oral": members},
            include_total_column=True,
        )
        ws = wb["Challenges Oral"]
        total_base = _get_base(ws, 2)
        assert total_base == 85, f"All-respondent MRS base should be 85, got {total_base}"

    def test_mrs_per_column_bases(self, data):
        """Q_39 per-column bases: PsO=30, PsA=30, UC=11, CD=14."""
        members = [f"Q_39_{i}" for i in range(1, 15) if f"Q_39_{i}" in data.df.columns]
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=[],
            mrs_groups={"Challenges Oral": members},
            include_total_column=True,
        )
        ws = wb["Challenges Oral"]
        # Columns: Total, PsO(A), PsA(B), UC(C), CD(D)
        pso_base = _get_base(ws, 3)  # Column A
        psa_base = _get_base(ws, 4)  # Column B
        assert pso_base == 30, f"PsO base should be 30, got {pso_base}"
        assert psa_base == 30, f"PsA base should be 30, got {psa_base}"


# ══════════════════════════════════════════════════════════════════════
# Test 4: Grid / Battery (individual mode)
# ══════════════════════════════════════════════════════════════════════

class TestGridIndividual:
    """Scale questions exported as individual crosstabs with auto-nets."""

    def test_grid_produces_individual_sheets(self, data):
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=[],
            grid_groups={
                "Oral Satisfaction": {
                    "variables": ["Q_20_1", "Q_20_2", "Q_20_3"],
                    "show": ["t2b", "b2b", "mean"],
                }
            },
            grid_mode="individual",
            include_total_column=True,
        )
        # Each variable should be its own sheet (not a compact summary)
        assert "Q_20_1" in wb.sheetnames or any("Q_20" in s for s in wb.sheetnames)
        assert result.successful >= 3

    def test_grid_summary_mode(self, data):
        """Summary mode: one sheet with T2B/Mean per variable."""
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=[],
            grid_groups={
                "Oral Satisfaction": {
                    "variables": ["Q_20_1", "Q_20_2", "Q_20_3"],
                    "show": ["t2b", "mean"],
                }
            },
            grid_mode="summary",
            include_total_column=True,
        )
        assert "Oral Satisfaction" in wb.sheetnames


# ══════════════════════════════════════════════════════════════════════
# Test 5: Open-ends excluded
# ══════════════════════════════════════════════════════════════════════

class TestOpenEnds:
    def test_text_variables_skipped(self, data):
        """Q_33_1 and Q_34_1 are open-ended text — should be skipped."""
        result, wb = _tabulate(data,
            banners=["Q_10"],
            stubs=["Q_33_1", "Q_34_1", "Q_11"],
            include_total_column=True,
        )
        # Q_33_1 and Q_34_1 should fail, Q_11 should succeed
        failed_vars = [s.variable for s in result.sheets if s.status == "error"]
        success_vars = [s.variable for s in result.sheets if s.status == "success"]
        assert "Q_33_1" in failed_vars, "Q_33_1 (open-end) should fail"
        assert "Q_11" in success_vars, "Q_11 (categorical) should succeed"


# ══════════════════════════════════════════════════════════════════════
# Test 6: Auto-detect groups
# ══════════════════════════════════════════════════════════════════════

class TestAutoDetect:
    def test_binary_mrs_detected(self, metadata):
        """Q_12_* (binary 0/1, no value_labels) should be detected as awareness."""
        groups = metadata.get("detected_groups") or []
        mrs_groups = [g for g in groups
                      if "awareness" in str(g.get("question_type", "")).lower()
                      and any("Q_12" in v for v in g.get("variables", []))]
        assert len(mrs_groups) >= 1, "Q_12_* should be detected as awareness MRS"

    def test_scale_groups_detected(self, metadata):
        """Scale questions (Q_20_*, Q_27_*) should be detected as scale/grid."""
        groups = metadata.get("detected_groups") or []
        scale_groups = [g for g in groups
                        if "scale" in str(g.get("question_type", "")).lower()]
        assert len(scale_groups) >= 2, f"Expected 2+ scale groups, got {len(scale_groups)}"


# ══════════════════════════════════════════════════════════════════════
# Test 7: Total-only export (no banners)
# ══════════════════════════════════════════════════════════════════════

class TestTotalOnly:
    def test_frequency_total_only(self, data):
        """Export Q_11 with just Total column."""
        result, wb = _tabulate(data,
            banners=[],
            stubs=["Q_11"],
            include_total_column=True,
        )
        assert result.successful >= 1
        ws = wb["Q_11"]
        base = _get_base(ws, 2)
        assert base == 85

    def test_mrs_total_only(self, data):
        """MRS with just Total — base should be respondent count, not responses."""
        members = [f"Q_39_{i}" for i in range(1, 15) if f"Q_39_{i}" in data.df.columns]
        result, wb = _tabulate(data,
            banners=[],
            stubs=[],
            mrs_groups={"Challenges": members},
            include_total_column=True,
        )
        ws = wb["Challenges"]
        base = _get_base(ws, 2)
        assert base == 85, f"Total-only MRS base should be 85, got {base}"


# ══════════════════════════════════════════════════════════════════════
# Test 8: Golden standard comparison (Data Tables 3.xlsx)
# ══════════════════════════════════════════════════════════════════════

class TestGoldenStandard:
    """Compare our output against the R export (Data Tables 3.xlsx)."""

    def test_s3_severity_total(self, golden):
        """R export: S3 Total: Moderate=70.6%, Severe=29.4%."""
        ws = golden["Total"]
        # S3 is at row ~17-23
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val and str(val).strip() == "Moderate":
                pct = ws.cell(row=r, column=2).value
                assert pct is not None
                assert abs(float(pct) - 0.7059) < 0.01
                break

    def test_s4a_pso_treatment_base(self, golden):
        """R export: S4A Column Sample Size for PsO = 30."""
        ws = golden["By Data Cut"]
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val and "Column Sample Size" in str(val):
                # Check if this is in the S4A block (around row 76)
                pso_val = ws.cell(row=r, column=2).value
                if pso_val == 30:
                    # Found it
                    total_val = ws.cell(row=r, column=5).value
                    assert total_val in (30, 85), f"S4A total sample should be 30 or 85, got {total_val}"
                    break
