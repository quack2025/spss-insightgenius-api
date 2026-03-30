"""Tests for POST /v1/tabulate endpoint — including Tier 1 features."""

import json
import io
from openpyxl import load_workbook


def test_tabulate_basic(client, auth_headers, test_sav_bytes):
    """Tabulate satisfaction x gender -> Excel with 1 sheet."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": ["satisfaction"],
        "significance_level": 0.95,
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp.headers["X-Stubs-Total"] == "1"
    assert resp.headers["X-Stubs-Success"] == "1"

    wb = load_workbook(io.BytesIO(resp.content))
    assert "Summary" in wb.sheetnames
    assert "satisfaction" in wb.sheetnames

    ws_ct = wb["satisfaction"]
    assert "satisfaction" in ws_ct.cell(row=1, column=1).value


def test_tabulate_all_stubs(client, auth_headers, test_sav_bytes):
    """Tabulate with stubs=_all_ auto-selects variables with value labels."""
    spec = json.dumps({"banner": "gender", "stubs": ["_all_"]})
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    total = int(resp.headers["X-Stubs-Total"])
    assert total >= 2


def test_tabulate_with_nets(client, auth_headers, test_sav_bytes):
    """Tabulate with net definitions (Top 2 Box)."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": ["satisfaction"],
        "nets": {"satisfaction": {"Top 2 Box": [4, 5], "Bottom 2 Box": [1, 2]}},
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["satisfaction"]
    nets_found = any(row[0].value == "Nets" for row in ws.iter_rows(min_col=1, max_col=1))
    assert nets_found, "Nets section not found in Excel"


def test_tabulate_with_weight(client, auth_headers, test_sav_bytes):
    """T1-4: Tabulate with weight shows dual bases."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": ["satisfaction"],
        "weight": "weight_var",
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["satisfaction"]
    # Check for dual bases
    found_weighted = False
    found_unweighted = False
    for row in ws.iter_rows(min_col=1, max_col=1, max_row=10):
        val = row[0].value or ""
        if "weighted" in val.lower():
            found_weighted = True
        if "unweighted" in val.lower():
            found_unweighted = True
    assert found_weighted, "Weighted base row not found"
    assert found_unweighted, "Unweighted base row not found"


def test_tabulate_means(client, auth_headers, test_sav_bytes):
    """T1-1: Means row with T-test."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": ["satisfaction"],
        "include_means": True,
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["satisfaction"]
    # Find "Mean" label
    mean_found = False
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value == "Mean":
            mean_found = True
            break
    assert mean_found, "Mean row not found in Excel"


def test_tabulate_multiple_banners(client, auth_headers, test_sav_bytes):
    """T1-2: Multiple banners side-by-side."""
    spec = json.dumps({
        "banners": ["gender", "age_group"],
        "stubs": ["satisfaction"],
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))

    # Summary should mention both banners
    ws_sum = wb["Summary"]
    banners_cell = None
    for row in ws_sum.iter_rows(min_col=1, max_col=2, max_row=15):
        if row[0].value == "Banners":
            banners_cell = row[1].value
            break
    assert banners_cell is not None
    assert "gender" in banners_cell
    assert "age_group" in banners_cell

    # Data sheet should have columns for both banners (2 gender + 3 age = 5 + Total = 6 data cols)
    ws = wb["satisfaction"]
    # Check letters row — should have A,B (gender) + C,D,E (age_group)
    letters = []
    for col in range(2, 10):
        val = ws.cell(row=5, column=col).value  # Letters row (row 5 with banner group header)
        if val and len(str(val)) <= 2 and str(val).isalpha():
            letters.append(str(val))
    assert len(letters) >= 5, f"Expected 5+ column letters, got {letters}"
    assert "A" in letters and "E" in letters


def test_tabulate_mrs(client, auth_headers, test_sav_bytes_with_mrs):
    """T1-3: MRS groups as crosstab rows."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": [],
        "mrs_groups": {"Brand Awareness": ["AWARE_A", "AWARE_B", "AWARE_C"]},
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test_mrs.sav", test_sav_bytes_with_mrs, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    assert int(resp.headers["X-Stubs-Success"]) >= 1
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Brand Awareness" in wb.sheetnames


def test_tabulate_total_first_column(client, auth_headers, test_sav_bytes):
    """T2-5: Total appears as first column before banners."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": ["satisfaction"],
        "include_total_column": True,
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["satisfaction"]
    # First data column (B) should be "Total"
    # Find the column value labels row
    for r in range(3, 8):
        val = ws.cell(row=r, column=2).value
        if val == "Total":
            break
    else:
        assert False, "Total not found as first column"


def test_tabulate_single_sheet(client, auth_headers, test_sav_bytes):
    """Single-sheet mode: all stubs stacked in one sheet."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": ["satisfaction", "age_group"],
        "output_mode": "single_sheet",
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Tabulation" in wb.sheetnames
    # Should NOT have individual sheets for each stub
    assert "satisfaction" not in wb.sheetnames
    assert "age_group" not in wb.sheetnames
    # The Tabulation sheet should contain both stubs
    ws = wb["Tabulation"]
    found_sat = False
    found_age = False
    for row in ws.iter_rows(min_col=1, max_col=1):
        val = str(row[0].value or "")
        if "satisfaction" in val.lower():
            found_sat = True
        if "age_group" in val.lower():
            found_age = True
    assert found_sat, "satisfaction stub not found in single sheet"
    assert found_age, "age_group stub not found in single sheet"


def test_tabulate_grid_summary(client, auth_headers, test_sav_bytes):
    """Grid/Battery summary: T2B + Mean for multiple scale variables in one sheet."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": [],
        "grid_mode": "summary",
        "grid_groups": {
            "Satisfaction Battery": {
                "variables": ["satisfaction", "recommend"],
                "show": ["t2b", "mean"]
            }
        },
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    assert int(resp.headers["X-Stubs-Success"]) >= 1
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Satisfaction Battery" in wb.sheetnames
    ws = wb["Satisfaction Battery"]
    # Should have "Top 2 Box %" and "Mean" metric headers
    found_t2b = False
    found_mean = False
    for row in ws.iter_rows(min_col=1, max_col=1):
        val = str(row[0].value or "")
        if "Top 2 Box" in val:
            found_t2b = True
        if val == "Mean":
            found_mean = True
    assert found_t2b, "T2B metric header not found"
    assert found_mean, "Mean metric header not found"


def test_tabulate_custom_groups(client, auth_headers, test_sav_bytes):
    """Custom groups: virtual banner columns defined by conditions."""
    spec = json.dumps({
        "banner": "gender",
        "stubs": ["satisfaction"],
        "custom_groups": [
            {"name": "Young Males", "conditions": [
                {"variable": "gender", "operator": "eq", "value": 1.0},
                {"variable": "age_group", "operator": "eq", "value": 1.0}
            ]},
            {"name": "Young Females", "conditions": [
                {"variable": "gender", "operator": "eq", "value": 2.0},
                {"variable": "age_group", "operator": "eq", "value": 1.0}
            ]},
        ],
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["satisfaction"]
    # Find "Young Males" in header rows
    found_custom = False
    for row in ws.iter_rows(min_row=1, max_row=8, min_col=2, max_col=10):
        for cell in row:
            if cell.value and "Young Males" in str(cell.value):
                found_custom = True
                break
    assert found_custom, "Custom group 'Young Males' not found in Excel headers"


def test_tabulate_custom_groups_only(client, auth_headers, test_sav_bytes):
    """Custom groups without regular banners."""
    spec = json.dumps({
        "stubs": ["satisfaction"],
        "custom_groups": [
            {"name": "Satisfied (4-5)", "conditions": [
                {"variable": "satisfaction", "operator": "gte", "value": 4.0}
            ]},
            {"name": "Neutral (3)", "conditions": [
                {"variable": "satisfaction", "operator": "eq", "value": 3.0}
            ]},
            {"name": "Dissatisfied (1-2)", "conditions": [
                {"variable": "satisfaction", "operator": "lte", "value": 2.0}
            ]},
        ],
    })
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    assert int(resp.headers["X-Stubs-Success"]) >= 1


def test_tabulate_total_only(client, auth_headers, test_sav_bytes):
    """Total-only export (no banners) should succeed with just the Total column."""
    spec = json.dumps({"stubs": ["satisfaction"]})
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 200
    assert resp.headers.get("content-type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_tabulate_invalid_banner(client, auth_headers, test_sav_bytes):
    spec = json.dumps({"banner": "NONEXISTENT", "stubs": ["satisfaction"]})
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 400
    assert resp.json()["detail"]["code"] == "VARIABLE_NOT_FOUND"


def test_tabulate_invalid_stub(client, auth_headers, test_sav_bytes):
    spec = json.dumps({"banner": "gender", "stubs": ["Q999"]})
    resp = client.post(
        "/v1/tabulate",
        headers=auth_headers,
        files={"file": ("test.sav", test_sav_bytes, "application/octet-stream")},
        data={"spec": spec},
    )
    assert resp.status_code == 400
    assert resp.json()["detail"]["code"] == "VARIABLE_NOT_FOUND"
