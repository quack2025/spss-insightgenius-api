"""Tabulation Excel builder — professional market research crosstab tables.

Generates an .xlsx workbook where each sheet is one stub variable crossed by
one or more banners, with significance letters, nets, means, and proper formatting.

Output mirrors what WinCross / PSPP / Quantum produce.

Tier 1 features:
- T1-1: Means row with T-test significance letters
- T1-2: Multiple banners side-by-side with grouped headers
- T1-3: MRS groups as crosstab rows
- T1-4: Dual bases (weighted + unweighted)
"""

import io
import logging
import math
import string
from dataclasses import dataclass, field
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats

logger = logging.getLogger(__name__)

# Excel sheet names cannot contain: \ / * ? : [ ]
_SHEET_NAME_BAD_CHARS = str.maketrans({':': '-', '/': '-', '\\': '-', '*': '', '?': '', '[': '(', ']': ')'})

def _sanitize_sheet_name(name: str) -> str:
    """Clean a string for use as an Excel sheet name (max 31 chars, no illegal chars)."""
    return name.translate(_SHEET_NAME_BAD_CHARS).strip()[:31]

# ── Styling ──────────────────────────────────────────────────────────────────

TITLE_FONT = Font(bold=True, size=12, color="1F2937")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
BANNER_GROUP_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
LETTER_HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
NET_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
MEAN_FILL = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
SIG_FONT = Font(bold=True, size=9, color="DC2626")
BASE_FONT = Font(bold=True, size=10, color="6B7280")
LABEL_FONT = Font(size=10, color="1F2937")
PCT_FONT = Font(size=10, color="374151")
COUNT_FONT = Font(size=9, color="9CA3AF")
MEAN_FONT = Font(bold=True, size=10, color="9A3412")
MEAN_SIG_FONT = Font(bold=True, size=9, color="DC2626")
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))
HEADER_BORDER = Border(bottom=Side(style="medium", color="1D4ED8"))
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


@dataclass
class CustomGroup:
    """A user-defined group (custom break / banner point).

    Conditions are AND-ed: all must be true for a row to be in this group.
    Example: {"name": "London Uber users", "conditions": [
        {"variable": "region", "operator": "eq", "value": 1},
        {"variable": "AWARE_UBER", "operator": "eq", "value": 1}
    ]}
    """
    name: str
    conditions: list[dict[str, Any]]  # [{"variable": str, "operator": str, "value": Any}]


# Supported operators for custom group conditions
_OPERATORS = {
    "eq": lambda s, v: s == v,
    "ne": lambda s, v: s != v,
    "gt": lambda s, v: s > v,
    "gte": lambda s, v: s >= v,
    "lt": lambda s, v: s < v,
    "lte": lambda s, v: s <= v,
    "in": lambda s, v: s.isin(v if isinstance(v, list) else [v]),
    "not_in": lambda s, v: ~s.isin(v if isinstance(v, list) else [v]),
}


def _apply_group_mask(df: pd.DataFrame, conditions: list[dict[str, Any]]) -> pd.Series:
    """Build a boolean mask from AND-ed conditions."""
    mask = pd.Series(True, index=df.index)
    for cond in conditions:
        var = cond.get("variable", "")
        op = cond.get("operator", "eq")
        val = cond.get("value")
        if var not in df.columns:
            continue
        op_fn = _OPERATORS.get(op, _OPERATORS["eq"])
        mask = mask & op_fn(df[var], val)
    return mask


@dataclass
class TabulateSpec:
    """Specification for a full tabulation run."""
    banner: str = ""  # Single banner (backwards compat)
    banners: list[str] | None = None  # Multiple banners (T1-2)
    custom_groups: list[dict[str, Any]] | None = None  # Custom breaks
    stubs: list[str] = field(default_factory=lambda: ["_all_"])
    weight: str | None = None
    significance_level: float = 0.95
    nets: dict[str, dict[str, list[Any]]] | None = None
    mrs_groups: dict[str, list[str]] | None = None
    grid_groups: dict[str, dict[str, Any]] | None = None  # Grid/Battery summaries
    # Format: {"Satisfaction": {"variables": ["sat_speed","sat_price",...], "show": ["t2b","b2b","mean"]}}
    grid_mode: str = "individual"  # "individual" (each var as crosstab) or "summary" (compact T2B/Mean table)
    include_means: bool = False
    include_total_column: bool = True  # Total as first column (T2-5)
    output_mode: str = "multi_sheet"  # "multi_sheet" or "single_sheet"
    show_counts: bool = True
    show_percentages: bool = True
    show_chi2: bool = True  # Chi-square p-value in Excel headers (Story #9)
    include_summary: bool = False  # AI executive summary as first sheet (Story #4)
    study_context: dict[str, Any] | None = None  # Optional context for summary
    filters: list[dict[str, Any]] | None = None  # Sub-population filters (Story #2)
    title: str = ""

    @property
    def resolved_banners(self) -> list[str]:
        if self.banners:
            return self.banners
        if self.banner:
            return [self.banner]
        return []

    @property
    def parsed_custom_groups(self) -> list[CustomGroup]:
        if not self.custom_groups:
            return []
        return [CustomGroup(name=g.get("name", f"Group {i+1}"), conditions=g.get("conditions", []))
                for i, g in enumerate(self.custom_groups)]


@dataclass
class BannerColumn:
    """A single column in the output table (one value of one banner variable)."""
    banner_var: str
    banner_label: str
    value: str  # The column value (e.g., "1.0")
    value_label: str  # The display label (e.g., "Male")
    letter: str  # The assigned letter (A, B, C, ...)
    banner_index: int  # Which banner group this belongs to


@dataclass
class GridGroupSpec:
    """A grid/battery summary group."""
    name: str
    variables: list[str]
    show: list[str]  # ["t2b", "b2b", "mean", "median"]


@dataclass
class SheetResult:
    """Result for one stub variable."""
    variable: str
    label: str | None
    status: str
    error: str | None = None
    crosstab_data: dict[str, Any] | None = None
    is_mrs: bool = False
    mrs_members: list[str] | None = None
    is_grid: bool = False
    grid_data: dict[str, Any] | None = None  # {"variables": [...], "metrics": {...}}


@dataclass
class TabulationResult:
    """Full tabulation result."""
    banners: list[str]
    banner_labels: list[str]
    total_stubs: int
    successful: int
    failed: int
    sheets: list[SheetResult] = field(default_factory=list)
    excel_bytes: bytes = b""
    banner_columns: list[BannerColumn] = field(default_factory=list)
    executive_summary: str = ""  # AI-generated summary (Story #4)


def build_tabulation(engine_cls: Any, data: Any, spec: TabulateSpec) -> TabulationResult:
    """Run all crosstabs and build the Excel workbook. CPU-bound — call via asyncio.to_thread()."""
    from services.quantipy_engine import QuantiProEngine

    df = data.df
    meta = data.meta

    # ── Apply filters / sub-population (Story #2) ──
    if spec.filters:
        mask = _apply_group_mask(df, spec.filters)
        df = df[mask].copy()
        data = type(data)(df=df, meta=meta, mrx_dataset=data.mrx_dataset, file_name=data.file_name)
        logger.info("Filters applied: %d → %d cases", len(data.df) + (~mask).sum(), len(df))

    col_labels = getattr(meta, "column_names_to_labels", {})
    banners = spec.resolved_banners

    # ── Insert Total as first banner if requested ──
    if spec.include_total_column:
        total_col_name = "_total_"
        df[total_col_name] = 1.0
        vvl = getattr(meta, "variable_value_labels", {})
        vvl[total_col_name] = {1.0: "Total"}
        cnl = getattr(meta, "column_names_to_labels", {})
        cnl[total_col_name] = "Total"
        banners = [total_col_name] + list(banners)

    # ── Build banner columns with continuous letter assignment ──
    banner_columns: list[BannerColumn] = []
    letter_idx = 0
    all_letters = list(string.ascii_uppercase) + [a + b for a in string.ascii_uppercase for b in string.ascii_uppercase]

    for b_idx, banner_var in enumerate(banners):
        if banner_var not in df.columns:
            continue
        vl = getattr(meta, "variable_value_labels", {}).get(banner_var, {})
        col_values = sorted(df[banner_var].dropna().unique())
        is_total = banner_var == "_total_" if spec.include_total_column else False
        for cv in col_values:
            banner_columns.append(BannerColumn(
                banner_var=banner_var,
                banner_label=col_labels.get(banner_var, banner_var),
                value=str(cv),
                value_label=vl.get(cv, str(cv)),
                letter="" if is_total else all_letters[letter_idx],
                banner_index=b_idx,
            ))
            if not is_total:
                letter_idx += 1

    # ── Custom groups as virtual banner columns ──
    custom_groups = spec.parsed_custom_groups
    custom_col_name = "_custom_group_"
    has_custom = len(custom_groups) > 0

    if has_custom:
        # Create a synthetic column in df where each custom group gets an integer code
        # Rows that match NO group get NaN (excluded from crosstab)
        custom_series = pd.Series(np.nan, index=df.index)
        custom_vl = {}
        for gi, cg in enumerate(custom_groups):
            mask = _apply_group_mask(df, cg.conditions)
            # Assign group code (1-based). Later groups overwrite earlier if overlapping.
            custom_series = custom_series.where(~mask, float(gi + 1))
            custom_vl[float(gi + 1)] = cg.name

        df[custom_col_name] = custom_series
        banners = list(banners) + [custom_col_name]

        cg_b_idx = len(banners) - 1
        for gi, cg in enumerate(custom_groups):
            banner_columns.append(BannerColumn(
                banner_var=custom_col_name,
                banner_label="Custom Groups",
                value=str(float(gi + 1)),
                value_label=cg.name,
                letter=all_letters[letter_idx],
                banner_index=cg_b_idx,
            ))
            letter_idx += 1

        # Patch meta so crosstab can find value labels
        if not hasattr(meta, "_custom_vl_patched"):
            vvl = getattr(meta, "variable_value_labels", {})
            vvl[custom_col_name] = custom_vl
            meta._custom_vl_patched = True
            cnl = getattr(meta, "column_names_to_labels", {})
            cnl[custom_col_name] = "Custom Groups"

    # ── Resolve stubs ──
    if spec.stubs == ["_all_"] or not spec.stubs:
        value_labels = getattr(meta, "variable_value_labels", {})
        banner_set = set(banners)
        # Collect variables that belong to MRS or Grid groups — they're exported as group sheets
        grouped_vars = set()
        for _name, members in (spec.mrs_groups or {}).items():
            grouped_vars.update(members)
        for _name, gspec in (spec.grid_groups or {}).items():
            grouped_vars.update(gspec.get("variables", []) if isinstance(gspec, dict) else [])
        stubs = [
            col for col in df.columns
            if col not in banner_set and col not in grouped_vars
            and col in value_labels and len(value_labels.get(col, {})) >= 2
        ]
        if not stubs:
            stubs = [col for col in df.columns
                     if col not in banner_set and col not in grouped_vars
                     and df[col].dtype.kind in ("i", "f")]
    else:
        stubs = spec.stubs

    result = TabulationResult(
        banners=banners,
        banner_labels=[col_labels.get(b, b) for b in banners],
        total_stubs=len(stubs) + len(spec.mrs_groups or {}) + len(spec.grid_groups or {}),
        successful=0,
        failed=0,
        banner_columns=banner_columns,
    )

    # ── Run crosstabs for each stub × each banner ──
    for stub in stubs:
        # Skip string/text variables (open-ends) — not tabulable
        if stub in df.columns and (df[stub].dtype == object or df[stub].dtype.kind in ('U', 'S', 'O')):
            result.sheets.append(SheetResult(variable=stub, label=col_labels.get(stub, stub), status="error", error="Text/open-end variable — not tabulable"))
            result.failed += 1
            continue
        try:
            # Run crosstab against FIRST banner (primary), store all banner results
            all_banner_results = {}
            for banner_var in banners:
                ct = QuantiProEngine.crosstab_with_significance(
                    data, row=stub, col=banner_var,
                    weight=spec.weight,
                    significance_level=spec.significance_level,
                )
                all_banner_results[banner_var] = ct

            sheet = SheetResult(
                variable=stub,
                label=col_labels.get(stub, stub),
                status="success",
                crosstab_data=all_banner_results,
            )
            result.successful += 1
        except Exception as e:
            sheet = SheetResult(variable=stub, label=col_labels.get(stub, stub), status="error", error=str(e))
            result.failed += 1
        result.sheets.append(sheet)

    # ── MRS groups (T1-3) ──
    for group_name, members in (spec.mrs_groups or {}).items():
        try:
            all_banner_results = {}
            for banner_var in banners:
                ct = _mrs_crosstab(data, members, banner_var, spec.weight, spec.significance_level)
                all_banner_results[banner_var] = ct
            group_label = col_labels.get(group_name, group_name)
            sheet = SheetResult(
                variable=group_name, label=group_label, status="success",
                crosstab_data=all_banner_results, is_mrs=True, mrs_members=members,
            )
            result.successful += 1
        except Exception as e:
            sheet = SheetResult(variable=group_name, label=str(group_name), status="error", error=str(e), is_mrs=True)
            result.failed += 1
        result.sheets.append(sheet)

    # ── Grid/Battery groups ──
    grid_mode = getattr(spec, "grid_mode", "individual")  # "individual" (default) or "summary"
    for grid_name, grid_cfg in (spec.grid_groups or {}).items():
        gvars = grid_cfg.get("variables", [])
        gshow = grid_cfg.get("show", ["t2b", "b2b", "mean"])

        if grid_mode == "summary":
            # Compact summary: one sheet with T2B/B2B/Mean rows per variable
            try:
                grid_result = _compute_grid_summary(data, gvars, banners, banner_columns, spec.weight, spec.significance_level, gshow)
                sheet = SheetResult(
                    variable=grid_name, label=grid_name, status="success",
                    is_grid=True, grid_data=grid_result,
                )
                result.successful += 1
            except Exception as e:
                sheet = SheetResult(variable=grid_name, label=grid_name, status="error", error=str(e), is_grid=True)
                result.failed += 1
            result.sheets.append(sheet)
        else:
            # Individual mode: each variable gets its own crosstab with auto-nets
            from services.quantipy_engine import _strip_common_label_prefix
            grid_labels = [col_labels.get(v, v) for v in gvars if v in df.columns]
            stripped = _strip_common_label_prefix(grid_labels)
            label_map = dict(zip([v for v in gvars if v in df.columns], stripped))

            for gvar in gvars:
                if gvar not in df.columns:
                    continue
                # Skip string/text variables (open-ends) — not tabulable
                if df[gvar].dtype == object or df[gvar].dtype.kind in ('U', 'S', 'O'):
                    continue
                try:
                    all_banner_results = {}
                    for banner_var in banners:
                        ct = engine_cls.crosstab_with_significance(
                            data, row=gvar, col=banner_var,
                            weight=spec.weight,
                            significance_level=spec.significance_level,
                        )
                        all_banner_results[banner_var] = ct
                    # Use stripped label for the variable
                    var_label = label_map.get(gvar, col_labels.get(gvar, gvar))
                    sheet = SheetResult(
                        variable=gvar, label=var_label, status="success",
                        crosstab_data=all_banner_results,
                    )
                    # Auto-add T2B/B2B nets for this variable
                    gvar_vl = getattr(meta, "variable_value_labels", {}).get(gvar, {})
                    if gvar_vl and len(gvar_vl) >= 4:
                        try:
                            keys = sorted([float(k) for k in gvar_vl.keys()])
                            if gvar not in (spec.nets or {}):
                                if not hasattr(spec, '_auto_grid_nets'):
                                    spec._auto_grid_nets = {}
                                spec._auto_grid_nets[gvar] = {
                                    "Top 2 Box": [int(keys[-2]), int(keys[-1])],
                                    "Bottom 2 Box": [int(keys[0]), int(keys[1])],
                                }
                        except (ValueError, TypeError):
                            pass
                    result.successful += 1
                except Exception as e:
                    sheet = SheetResult(variable=gvar, label=label_map.get(gvar, gvar), status="error", error=str(e))
                    result.failed += 1
                result.sheets.append(sheet)

    result.excel_bytes = _build_excel(result, spec, data)
    return result


def _mrs_crosstab(
    data: Any, members: list[str], col: str, weight: str | None, sig_level: float,
) -> dict[str, Any]:
    """Crosstab for a Multiple Response Set.

    Base = respondents who answered AT LEAST ONE member (not NaN in all members).
    This correctly handles conditional/show-if questions where only a subgroup was asked.
    """
    df = data.df
    meta = data.meta
    col_labels = getattr(meta, "column_names_to_labels", {})
    col_vl = getattr(meta, "variable_value_labels", {}).get(col, {})

    valid_members = [m for m in members if m in df.columns]
    # Start with rows that have a valid banner value
    valid = df[[col] + valid_members + ([weight] if weight and weight in df.columns else [])].dropna(subset=[col])

    # Key fix: filter to respondents who answered AT LEAST ONE MRS member
    # (not NaN in all members). This handles conditional/show-if questions.
    mrs_answered = valid[valid_members].notna().any(axis=1)
    valid = valid.loc[mrs_answered]

    w = valid[weight] if weight and weight in df.columns else None

    col_values = sorted(valid[col].dropna().unique())
    letters = list(string.ascii_uppercase[:len(col_values)])
    col_letter_map = {str(v): letters[i] for i, v in enumerate(col_values)}
    alpha = 1 - sig_level

    # Strip common label prefix for MRS (e.g., "What treatments...? -Otezla" → "Otezla")
    from services.quantipy_engine import _strip_common_label_prefix
    raw_labels = [col_labels.get(m, m) for m in valid_members]
    stripped_labels = _strip_common_label_prefix(raw_labels)
    _member_label_map = dict(zip(valid_members, stripped_labels))

    # Pre-compute per-column bases (only respondents who answered the MRS)
    col_bases = {}
    col_subsets = {}
    for i, cv in enumerate(col_values):
        mask = valid[col] == cv
        subset = valid.loc[mask]
        col_subsets[cv] = subset
        if w is not None:
            col_bases[cv] = float(w.loc[subset.index].sum())
        else:
            col_bases[cv] = len(subset)

    table = []
    for member in valid_members:
        member_label = _member_label_map.get(member, col_labels.get(member, member))
        row_data = {"row_value": member, "row_label": member_label}

        for i, cv in enumerate(col_values):
            subset = col_subsets[cv]
            base = col_bases[cv]
            # Count respondents who selected this option (value == 1 or value > 0)
            selected = (subset[member].fillna(0) > 0)
            if w is not None:
                w_sub = w.loc[subset.index]
                count = float((selected * w_sub).sum())
            else:
                count = int(selected.sum())
            pct = round(count / base * 100, 1) if base > 0 else 0

            # Sig testing vs other columns
            sig_letters = []
            for j, ocv in enumerate(col_values):
                if i == j:
                    continue
                o_subset = col_subsets[ocv]
                o_base = col_bases[ocv]
                o_selected = (o_subset[member].fillna(0) > 0)
                if w is not None:
                    o_count = float((o_selected * w.loc[o_subset.index]).sum())
                else:
                    o_count = int(o_selected.sum())
                try:
                    p1 = count / base if base > 0 else 0
                    p2 = o_count / o_base if o_base > 0 else 0
                    p_pool = (count + o_count) / (base + o_base) if (base + o_base) > 0 else 0
                    se = np.sqrt(p_pool * (1 - p_pool) * (1 / max(base, 1) + 1 / max(o_base, 1)))
                    if se > 0:
                        z = (p1 - p2) / se
                        p_val = 2 * stats.norm.sf(abs(z))
                        if p_val < alpha and p1 > p2:
                            sig_letters.append(letters[j])
                except Exception:
                    pass

            row_data[str(cv)] = {
                "count": count, "percentage": pct,
                "column_letter": letters[i],
                "significance_letters": sorted(set(sig_letters)),
            }
        table.append(row_data)

    return {
        "row_variable": "MRS",
        "col_variable": col,
        "total_responses": len(valid),
        "table": table,
        "col_labels": col_letter_map,
        "col_value_labels": {str(v): col_vl.get(v, str(v)) for v in col_values},
        "significance_level": sig_level,
        "significant_pairs": [],
    }


# ── Means computation with T-test (T1-1) ──

def _compute_grid_summary(
    data: Any, variables: list[str], banners: list[str],
    banner_columns: list, weight: str | None, sig_level: float, show: list[str],
) -> dict[str, Any]:
    """Compute T2B, B2B, Mean per variable per banner column for a grid/battery summary."""
    df = data.df
    meta = data.meta
    col_labels = getattr(meta, "column_names_to_labels", {})
    value_labels_all = getattr(meta, "variable_value_labels", {})
    alpha = 1 - sig_level

    rows = []  # Each row: {"variable": str, "label": str, "metric": str, "columns": {letter: {value, sig_letters}}}
    w = df[weight] if weight and weight in df.columns else None

    for var in variables:
        if var not in df.columns or not pd.api.types.is_numeric_dtype(df[var]):
            continue

        series = df[var].dropna()
        vl = value_labels_all.get(var, {})
        var_label = col_labels.get(var, var)

        # Detect scale endpoints for T2B/B2B
        if vl:
            scale_vals = sorted([k for k in vl.keys() if isinstance(k, (int, float))])
        else:
            scale_vals = sorted(series.unique())
        if len(scale_vals) < 3:
            continue

        top2 = scale_vals[-2:]
        bot2 = scale_vals[:2]

        for metric in show:
            row = {"variable": var, "label": var_label, "metric": metric, "columns": {}}

            # Compute metric per banner column
            for bc in banner_columns:
                bvar = bc.banner_var
                if bvar not in df.columns:
                    continue
                try:
                    bval = float(bc.value) if bc.value.replace('.', '', 1).replace('-', '', 1).isdigit() else bc.value
                except (ValueError, AttributeError):
                    bval = bc.value
                col_mask = df[bvar] == bval
                subset = df.loc[col_mask & series.notna()]
                sub_series = series[subset.index]

                if len(sub_series) == 0:
                    row["columns"][bc.letter] = {"value": None, "sig_letters": []}
                    continue

                sub_w = w[subset.index] if w is not None else None

                if metric == "t2b":
                    if sub_w is not None:
                        val = float((sub_series.isin(top2) * sub_w).sum()) / float(sub_w.sum()) * 100
                    else:
                        val = float(sub_series.isin(top2).sum()) / len(sub_series) * 100
                    row["columns"][bc.letter] = {"value": round(val, 1), "sig_letters": []}

                elif metric == "b2b":
                    if sub_w is not None:
                        val = float((sub_series.isin(bot2) * sub_w).sum()) / float(sub_w.sum()) * 100
                    else:
                        val = float(sub_series.isin(bot2).sum()) / len(sub_series) * 100
                    row["columns"][bc.letter] = {"value": round(val, 1), "sig_letters": []}

                elif metric == "mean":
                    if sub_w is not None:
                        val = float(np.average(sub_series, weights=sub_w))
                    else:
                        val = float(sub_series.mean())
                    row["columns"][bc.letter] = {"value": round(val, 2), "sig_letters": []}

                elif metric == "median":
                    val = float(sub_series.median())
                    row["columns"][bc.letter] = {"value": round(val, 2), "sig_letters": []}

            # Sig testing for means (T-test between columns within same banner group)
            if metric == "mean":
                _add_mean_sig_to_grid_row(row, df, series, banner_columns, w, alpha)
            elif metric in ("t2b", "b2b"):
                target_vals = top2 if metric == "t2b" else bot2
                _add_prop_sig_to_grid_row(row, df, series, target_vals, banner_columns, w, alpha)

            # Total column
            if w is not None:
                valid_w = w[series.notna().reindex(w.index, fill_value=False)]
            if metric == "t2b":
                total = float(series.isin(top2).sum()) / len(series) * 100 if len(series) > 0 else 0
            elif metric == "b2b":
                total = float(series.isin(bot2).sum()) / len(series) * 100 if len(series) > 0 else 0
            elif metric == "mean":
                total = float(np.average(series, weights=w[series.index]) if w is not None else series.mean())
            elif metric == "median":
                total = float(series.median())
            else:
                total = 0
            row["total"] = round(total, 2)

            rows.append(row)

    return {"rows": rows}


def _add_mean_sig_to_grid_row(row, df, series, banner_columns, w, alpha):
    """Add T-test sig letters to mean metric in grid row."""
    # Group columns by banner_index for within-group testing
    groups = {}
    for bc in banner_columns:
        groups.setdefault(bc.banner_index, []).append(bc)

    for bcs in groups.values():
        for i, bc_i in enumerate(bcs):
            for j, bc_j in enumerate(bcs):
                if i == j:
                    continue
                try:
                    bval_i = float(bc_i.value) if bc_i.value.replace('.','',1).replace('-','',1).isdigit() else bc_i.value
                    bval_j = float(bc_j.value) if bc_j.value.replace('.','',1).replace('-','',1).isdigit() else bc_j.value
                    mask_i = df[bc_i.banner_var] == bval_i
                    mask_j = df[bc_j.banner_var] == bval_j
                    vals_i = series[mask_i & series.notna()]
                    vals_j = series[mask_j & series.notna()]
                    if len(vals_i) < 2 or len(vals_j) < 2:
                        continue
                    t_stat, p_val = stats.ttest_ind(vals_i, vals_j, equal_var=False)
                    if p_val < alpha and vals_i.mean() > vals_j.mean():
                        if bc_i.letter in row["columns"]:
                            row["columns"][bc_i.letter]["sig_letters"].append(bc_j.letter)
                except Exception:
                    pass


def _add_prop_sig_to_grid_row(row, df, series, target_vals, banner_columns, w, alpha):
    """Add z-test sig letters to proportion metric (T2B/B2B) in grid row."""
    groups = {}
    for bc in banner_columns:
        groups.setdefault(bc.banner_index, []).append(bc)

    for bcs in groups.values():
        for i, bc_i in enumerate(bcs):
            for j, bc_j in enumerate(bcs):
                if i == j:
                    continue
                try:
                    bval_i = float(bc_i.value) if bc_i.value.replace('.','',1).replace('-','',1).isdigit() else bc_i.value
                    bval_j = float(bc_j.value) if bc_j.value.replace('.','',1).replace('-','',1).isdigit() else bc_j.value
                    mask_i = df[bc_i.banner_var] == bval_i
                    mask_j = df[bc_j.banner_var] == bval_j
                    sub_i = series[mask_i & series.notna()]
                    sub_j = series[mask_j & series.notna()]
                    n_i, n_j = len(sub_i), len(sub_j)
                    if n_i < 2 or n_j < 2:
                        continue
                    c_i = int(sub_i.isin(target_vals).sum())
                    c_j = int(sub_j.isin(target_vals).sum())
                    from statsmodels.stats.proportion import proportions_ztest
                    z_stat, p_val = proportions_ztest([c_i, c_j], [n_i, n_j], alternative="two-sided")
                    p_i = c_i / n_i
                    p_j = c_j / n_j
                    if p_val < alpha and p_i > p_j:
                        if bc_i.letter in row["columns"]:
                            row["columns"][bc_i.letter]["sig_letters"].append(bc_j.letter)
                except Exception:
                    pass


def _compute_means_by_column(
    df: pd.DataFrame, stub: str, banner_var: str, col_values: list, weight: str | None, alpha: float,
) -> dict[str, Any]:
    """Compute mean per banner column with independent T-test significance letters."""
    results = {}
    series = df[stub]
    if not pd.api.types.is_numeric_dtype(series):
        return {}

    w = df[weight] if weight and weight in df.columns else None
    letters = list(string.ascii_uppercase[:len(col_values)])

    col_stats = {}
    for i, cv in enumerate(col_values):
        mask = (df[banner_var] == cv) & series.notna()
        vals = series[mask]
        if len(vals) < 2:
            col_stats[str(cv)] = {"mean": None, "std": None, "n": len(vals), "letter": letters[i], "values": vals}
            continue
        if w is not None:
            wv = w[mask]
            wmean = float(np.average(vals, weights=wv))
            # Weighted std dev
            wvar = float(np.average((vals - wmean) ** 2, weights=wv))
            wstd = float(np.sqrt(wvar))
            col_stats[str(cv)] = {"mean": wmean, "std": wstd, "n": len(vals), "letter": letters[i], "values": vals, "weights": wv}
        else:
            col_stats[str(cv)] = {"mean": float(vals.mean()), "std": float(vals.std()), "n": len(vals), "letter": letters[i], "values": vals}

    # T-test between each pair
    for cv_str, st in col_stats.items():
        if st["mean"] is None:
            results[cv_str] = {"mean": None, "std": None, "n": st["n"], "sig_letters": []}
            continue
        sig_letters = []
        for ocv_str, ost in col_stats.items():
            if cv_str == ocv_str or ost["mean"] is None:
                continue
            try:
                t_stat, p_val = stats.ttest_ind(st["values"], ost["values"], equal_var=False)
                if p_val < alpha and st["mean"] > ost["mean"]:
                    sig_letters.append(ost["letter"])
            except Exception:
                pass
        results[cv_str] = {
            "mean": round(st["mean"], 2),
            "std": round(st["std"], 2),
            "n": st["n"],
            "sig_letters": sorted(set(sig_letters)),
        }
    return results


# ── Excel builder ──

def _build_excel(result: TabulationResult, spec: TabulateSpec, data: Any) -> bytes:
    wb = Workbook()

    # Executive Summary sheet (Story #4) — before Summary if available
    if result.executive_summary:
        ws_exec = wb.active
        ws_exec.title = "Executive Summary"
        ws_exec.cell(row=1, column=1, value="Executive Summary").font = TITLE_FONT
        ws_exec.cell(row=2, column=1, value=f"Generated by AI from {result.total_stubs} analysis tables").font = Font(italic=True, size=9, color="6B7280")
        row = 4
        for line in result.executive_summary.split("\n"):
            ws_exec.cell(row=row, column=1, value=line)
            if line.startswith("- ") or line.startswith("• "):
                ws_exec.cell(row=row, column=1).font = Font(size=10)
            elif line.startswith("#"):
                ws_exec.cell(row=row, column=1).font = Font(bold=True, size=11)
            row += 1
        ws_exec.column_dimensions["A"].width = 100
        ws_summary = wb.create_sheet(title="Summary")
    else:
        ws_summary = wb.active
        ws_summary.title = "Summary"

    _write_summary_sheet(ws_summary, result, spec, data)

    if spec.output_mode == "single_sheet":
        ws = wb.create_sheet(title="Tabulation")
        _write_single_sheet(ws, result, spec, data)
    else:
        for sheet_result in result.sheets:
            if sheet_result.status != "success":
                continue
            sheet_name = _sanitize_sheet_name(sheet_result.variable)
            existing = [ws.title for ws in wb.worksheets]
            if sheet_name in existing:
                sheet_name = sheet_name[:28] + "_" + str(existing.count(sheet_name))
            ws = wb.create_sheet(title=sheet_name)
            if sheet_result.is_grid and sheet_result.grid_data:
                _write_grid_sheet(ws, sheet_result, spec, result.banner_columns)
            elif sheet_result.crosstab_data:
                _write_crosstab_sheet(ws, sheet_result, spec, data, result.banner_columns)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_summary_sheet(ws, result: TabulationResult, spec: TabulateSpec, data: Any):
    meta = data.meta
    col_labels = getattr(meta, "column_names_to_labels", {})

    title = spec.title or "Tabulation Report"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=16, color="1F2937")
    ws.merge_cells("A1:F1")

    r = 3
    banners_str = " + ".join(f"{b} ({col_labels.get(b, '')})" for b in result.banners)
    info = [
        ("File", data.file_name),
        ("Cases", len(data.df)),
        ("Banners", banners_str),
        ("Weight", spec.weight or "(none)"),
        ("Significance Level", f"{spec.significance_level:.0%}"),
        ("Means with T-test", "Yes" if spec.include_means else "No"),
        ("Total Stubs", result.total_stubs),
        ("Successful", result.successful),
        ("Failed", result.failed),
    ]
    for label, value in info:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=str(value)).font = Font(size=10)
        r += 1

    # Column legend
    r += 1
    ws.cell(row=r, column=1, value="Column Legend").font = Font(bold=True, size=11)
    r += 1
    headers = ["Letter", "Banner", "Value", "Label"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    r += 1
    for bc in result.banner_columns:
        ws.cell(row=r, column=1, value=bc.letter).font = Font(bold=True, size=10)
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2, value=bc.banner_label)
        ws.cell(row=r, column=3, value=bc.value)
        ws.cell(row=r, column=4, value=bc.value_label)
        r += 1

    # Stub index
    r += 1
    ws.cell(row=r, column=1, value="Stub Index").font = Font(bold=True, size=11)
    r += 1
    idx_headers = ["#", "Variable", "Label", "Type", "Status"]
    for c, h in enumerate(idx_headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    r += 1
    for i, sheet in enumerate(result.sheets, 1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=sheet.variable)
        ws.cell(row=r, column=3, value=sheet.label or "")
        ws.cell(row=r, column=4, value="MRS" if sheet.is_mrs else "Standard")
        ws.cell(row=r, column=5, value=sheet.status)
        if sheet.status == "error":
            ws.cell(row=r, column=5).font = Font(color="DC2626")
        r += 1

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = [18, 25, 12, 45, 12, 20]["ABCDEF".index(col_letter)]


def _write_crosstab_sheet(ws, sheet_result: SheetResult, spec: TabulateSpec, data: Any, banner_columns: list[BannerColumn]):
    """Write one crosstab table with multiple banners side-by-side."""
    all_ct = sheet_result.crosstab_data  # dict[banner_var -> crosstab_result]
    if not all_ct:
        return

    banners = spec.resolved_banners
    n_banner_cols = len(banner_columns)
    last_data_col = 1 + n_banner_cols  # Column A = labels, then banner cols

    # Row 1: Title
    title = f"{sheet_result.variable}"
    if sheet_result.label:
        title += f": {sheet_result.label}"
    if sheet_result.is_mrs:
        title += " (Multiple Response)"
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_data_col)

    # Row 2: Sig note + Chi-square p-value (Story #9)
    sig_note = f"Significance: {spec.significance_level:.0%} confidence"
    if spec.weight:
        sig_note += f" | Weighted by: {spec.weight}"
    if len(banners) > 1:
        sig_note += f" | Sig testing within each banner group"

    # Add chi-square for the first banner's crosstab (most relevant)
    if spec.show_chi2:
        first_banner = banners[0] if banners else None
        first_ct = all_ct.get(first_banner, {}) if first_banner else {}
        chi2_p = first_ct.get("chi2_pvalue")
        if chi2_p is not None and isinstance(chi2_p, (int, float)):
            stars = "***" if chi2_p < 0.001 else "**" if chi2_p < 0.01 else "*" if chi2_p < 0.05 else "ns"
            sig_note += f" | Chi²: p={chi2_p:.4f} {stars}"

    ws.cell(row=2, column=1, value=sig_note).font = Font(italic=True, size=9, color="6B7280")

    # Row 3: Banner group headers (merged cells per banner)
    if len(banners) > 1:
        col_idx = 2
        for b_idx, banner_var in enumerate(banners):
            group_cols = [bc for bc in banner_columns if bc.banner_index == b_idx]
            if group_cols:
                cell = ws.cell(row=3, column=col_idx, value=group_cols[0].banner_label)
                cell.font = Font(bold=True, size=10, color="FFFFFF")
                cell.fill = BANNER_GROUP_FILL
                cell.alignment = CENTER
                if len(group_cols) > 1:
                    ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx + len(group_cols) - 1)
                col_idx += len(group_cols)
        header_row_offset = 1
    else:
        header_row_offset = 0

    row_labels = 3 + header_row_offset  # Column value labels
    row_letters = 4 + header_row_offset  # Letters
    row_base_w = 5 + header_row_offset  # Base (weighted or only)
    has_dual_base = spec.weight is not None
    row_base_uw = (6 + header_row_offset) if has_dual_base else None
    data_start = (7 if has_dual_base else 6) + header_row_offset

    # Row: Column value labels
    for i, bc in enumerate(banner_columns):
        is_total = bc.banner_var == "_total_"
        cell = ws.cell(row=row_labels, column=2 + i, value=bc.value_label)
        cell.font = HEADER_FONT
        cell.fill = TOTAL_FILL if is_total else HEADER_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER

    # Row: Letters
    for i, bc in enumerate(banner_columns):
        cell = ws.cell(row=row_letters, column=2 + i, value=bc.letter or "")
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = LETTER_HEADER_FILL
        cell.alignment = CENTER

    # Row: Base (weighted) — compute from data
    base_label = "Base (weighted)" if has_dual_base else "Base (N)"
    ws.cell(row=row_base_w, column=1, value=base_label).font = BASE_FONT
    col_bases = {}  # column index -> weighted count
    col_bases_uw = {}

    for i, bc in enumerate(banner_columns):
        ct = all_ct.get(bc.banner_var, {})
        table = ct.get("table", [])
        w_total = sum(
            row_data.get(bc.value, {}).get("count", 0)
            for row_data in table if isinstance(row_data.get(bc.value), dict)
        )
        col_bases[i] = w_total

        if has_dual_base:
            try:
                bval = float(bc.value) if bc.value.replace('.', '', 1).replace('-', '', 1).isdigit() else bc.value
            except (ValueError, AttributeError):
                bval = bc.value
            mask = data.df[bc.banner_var] == bval
            col_bases_uw[i] = int(mask.sum())

        is_total = bc.banner_var == "_total_"
        cell = ws.cell(row=row_base_w, column=2 + i, value=int(round(w_total)))
        cell.font = BASE_FONT
        cell.alignment = CENTER
        cell.fill = TOTAL_FILL if is_total else PatternFill()

    # Row: Base (unweighted) — T1-4
    if has_dual_base:
        ws.cell(row=row_base_uw, column=1, value="Base (unweighted)").font = Font(bold=True, size=9, color="9CA3AF")
        for i, bc in enumerate(banner_columns):
            uw = col_bases_uw.get(i, 0)
            cell = ws.cell(row=row_base_uw, column=2 + i, value=uw)
            cell.font = Font(size=9, color="9CA3AF")
            cell.alignment = CENTER

    # ── Data rows ──
    # Collect all unique row values/labels across all banner crosstabs
    all_rows = []
    seen_rv = set()
    for banner_var in banners:
        ct = all_ct.get(banner_var, {})
        for row_data in ct.get("table", []):
            rv = row_data.get("row_value")
            if rv not in seen_rv:
                seen_rv.add(rv)
                all_rows.append({"row_value": rv, "row_label": row_data.get("row_label", str(rv))})

    current_row = data_start
    for row_info in all_rows:
        rv = row_info["row_value"]
        ws.cell(row=current_row, column=1, value=row_info["row_label"]).font = LABEL_FONT
        ws.cell(row=current_row, column=1).alignment = LEFT

        row_total_count = 0
        for i, bc in enumerate(banner_columns):
            ct = all_ct.get(bc.banner_var, {})
            # Find this row in the crosstab
            cell_data = None
            for rd in ct.get("table", []):
                if rd.get("row_value") == rv:
                    cell_data = rd.get(bc.value, {})
                    break
            if not isinstance(cell_data, dict):
                ws.cell(row=current_row, column=2 + i, value="-").font = COUNT_FONT
                ws.cell(row=current_row, column=2 + i).alignment = CENTER
                continue

            count = cell_data.get("count", 0)
            pct = cell_data.get("percentage", 0)
            orig_sig = cell_data.get("significance_letters", [])
            row_total_count += count

            # Remap sig letters from per-banner letters to global letters
            ct_col_labels = ct.get("col_labels", {})
            # Build reverse map: original letter -> value
            orig_letter_to_value = {v: k for k, v in ct_col_labels.items()}
            # Map to global letters
            global_sig = []
            for orig_letter in orig_sig:
                orig_value = orig_letter_to_value.get(orig_letter)
                if orig_value:
                    for gbc in banner_columns:
                        if gbc.banner_var == bc.banner_var and gbc.value == orig_value:
                            global_sig.append(gbc.letter)
                            break

            col_idx = 2 + i
            is_total_col = bc.banner_var == "_total_"
            if spec.show_percentages:
                pct_str = f"{pct:.1f}%"
                if global_sig and not is_total_col:
                    pct_str += " " + "".join(global_sig)
                    cell = ws.cell(row=current_row, column=col_idx, value=pct_str)
                    cell.font = SIG_FONT
                else:
                    cell = ws.cell(row=current_row, column=col_idx, value=pct_str)
                    cell.font = PCT_FONT
                cell.alignment = CENTER
                if is_total_col:
                    cell.fill = TOTAL_FILL

        for c in range(1, last_data_col + 1):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

    # ── Nets ──
    var_nets = (spec.nets or {}).get(sheet_result.variable, {})
    if var_nets:
        current_row += 1
        ws.cell(row=current_row, column=1, value="Nets").font = Font(bold=True, size=10, color="166534")
        current_row += 1
        for net_name, net_values in var_nets.items():
            ws.cell(row=current_row, column=1, value=net_name).font = Font(bold=True, size=10, color="166534")
            ws.cell(row=current_row, column=1).fill = NET_FILL

            for i, bc in enumerate(banner_columns):
                ct = all_ct.get(bc.banner_var, {})
                net_count = 0
                col_base = col_bases.get(i, 0)
                for rd in ct.get("table", []):
                    rv = rd.get("row_value")
                    if rv in net_values or (isinstance(rv, float) and int(rv) in net_values):
                        cd = rd.get(bc.value, {})
                        if isinstance(cd, dict):
                            net_count += cd.get("count", 0)
                net_pct = round(net_count / col_base * 100, 1) if col_base > 0 else 0
                cell = ws.cell(row=current_row, column=2 + i, value=f"{net_pct:.1f}%")
                cell.font = Font(size=10, color="166534")
                cell.fill = NET_FILL
                cell.alignment = CENTER
            current_row += 1

    # ── Means with T-test (T1-1) ──
    if spec.include_means and not sheet_result.is_mrs:
        stub_var = sheet_result.variable
        if pd.api.types.is_numeric_dtype(data.df[stub_var]):
            current_row += 1
            ws.cell(row=current_row, column=1, value="Statistics").font = Font(bold=True, size=10, color="9A3412")
            current_row += 1
            alpha = 1 - spec.significance_level

            # Compute means per banner group
            for b_idx, banner_var in enumerate(banners):
                group_cols = [bc for bc in banner_columns if bc.banner_index == b_idx]
                col_vals = [float(bc.value) if bc.value.replace('.', '', 1).replace('-', '', 1).isdigit() else bc.value for bc in group_cols]
                means_result = _compute_means_by_column(data.df, stub_var, banner_var, col_vals, spec.weight, alpha)

                # Mean row
                ws.cell(row=current_row, column=1, value="Mean").font = MEAN_FONT
                ws.cell(row=current_row, column=1).fill = MEAN_FILL
                for bc in group_cols:
                    col_i = banner_columns.index(bc)
                    mr = means_result.get(str(float(bc.value) if bc.value.replace('.', '', 1).replace('-', '', 1).isdigit() else bc.value), {})
                    m = mr.get("mean")
                    if m is None:
                        cell = ws.cell(row=current_row, column=2 + col_i, value="-")
                    else:
                        sig = mr.get("sig_letters", [])
                        # Remap to global letters
                        global_sig = []
                        for sl in sig:
                            for gbc in group_cols:
                                if gbc.letter == sl or (hasattr(gbc, '_local_letter') and gbc._local_letter == sl):
                                    global_sig.append(gbc.letter)
                                    break
                            else:
                                # Try matching by index
                                idx = ord(sl) - ord('A')
                                if 0 <= idx < len(group_cols):
                                    global_sig.append(group_cols[idx].letter)

                        mean_str = f"{m:.2f}"
                        if global_sig:
                            mean_str += " " + "".join(global_sig)
                            cell = ws.cell(row=current_row, column=2 + col_i, value=mean_str)
                            cell.font = MEAN_SIG_FONT
                        else:
                            cell = ws.cell(row=current_row, column=2 + col_i, value=mean_str)
                            cell.font = MEAN_FONT
                    cell.fill = MEAN_FILL
                    cell.alignment = CENTER

            current_row += 1

            # Std Dev row
            ws.cell(row=current_row, column=1, value="Std Dev").font = Font(size=9, color="9A3412")
            ws.cell(row=current_row, column=1).fill = MEAN_FILL
            for bc in banner_columns:
                col_i = banner_columns.index(bc)
                cv_float = float(bc.value) if bc.value.replace('.', '', 1).replace('-', '', 1).isdigit() else bc.value
                means_result = _compute_means_by_column(data.df, stub_var, bc.banner_var, [cv_float], spec.weight, alpha)
                mr = means_result.get(str(cv_float), {})
                std = mr.get("std")
                cell = ws.cell(row=current_row, column=2 + col_i, value=f"{std:.2f}" if std is not None else "-")
                cell.font = Font(size=9, color="9A3412")
                cell.fill = MEAN_FILL
                cell.alignment = CENTER
            current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 40
    for i in range(n_banner_cols):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16
    ws.freeze_panes = f"B{data_start}"


def _write_grid_sheet(ws, sheet_result: SheetResult, spec: TabulateSpec, banner_columns: list[BannerColumn]):
    """Write a grid/battery summary sheet — compact view of multiple variables.

    Layout:
    Row 1: Title
    Row 2: Sig note
    Row 3: Banner group headers (if multi-banner)
    Row 4: Column value labels
    Row 5: Column letters
    Row 6+: One row per variable per metric, grouped by metric
    """
    grid = sheet_result.grid_data
    if not grid:
        return

    rows = grid.get("rows", [])
    banners = spec.resolved_banners
    n_cols = len(banner_columns)
    total_col = 2 + n_cols

    METRIC_LABELS = {"t2b": "Top 2 Box %", "b2b": "Bottom 2 Box %", "mean": "Mean", "median": "Median"}
    METRIC_FILL = {
        "t2b": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "b2b": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "mean": MEAN_FILL,
        "median": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    }
    METRIC_FONT = {
        "t2b": Font(size=10, color="166534"),
        "b2b": Font(size=10, color="991B1B"),
        "mean": MEAN_FONT,
        "median": Font(size=10, color="5B21B6"),
    }

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"{sheet_result.label or sheet_result.variable} (Grid Summary)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)

    # Row 2: Sig note
    sig_note = f"Significance: {spec.significance_level:.0%} confidence"
    if spec.weight:
        sig_note += f" | Weighted by: {spec.weight}"
    ws.cell(row=2, column=1, value=sig_note).font = Font(italic=True, size=9, color="6B7280")

    # Row 3: Banner group headers
    has_multi = len(banners) > 1
    offset = 1 if has_multi else 0
    if has_multi:
        col_idx = 2
        for b_idx, bvar in enumerate(banners):
            group_cols = [bc for bc in banner_columns if bc.banner_index == b_idx]
            if group_cols:
                cell = ws.cell(row=3, column=col_idx, value=group_cols[0].banner_label)
                cell.font = Font(bold=True, size=10, color="FFFFFF")
                cell.fill = BANNER_GROUP_FILL
                cell.alignment = CENTER
                if len(group_cols) > 1:
                    ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx + len(group_cols) - 1)
                col_idx += len(group_cols)

    # Column labels
    label_row = 3 + offset
    for i, bc in enumerate(banner_columns):
        cell = ws.cell(row=label_row, column=2 + i, value=bc.value_label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    cell = ws.cell(row=label_row, column=total_col, value="Total")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER

    # Letters
    letter_row = label_row + 1
    for i, bc in enumerate(banner_columns):
        cell = ws.cell(row=letter_row, column=2 + i, value=bc.letter)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = LETTER_HEADER_FILL
        cell.alignment = CENTER

    # Data rows grouped by metric
    current_row = letter_row + 1
    metrics_seen = []
    for r in rows:
        if r["metric"] not in metrics_seen:
            metrics_seen.append(r["metric"])

    for metric in metrics_seen:
        # Metric header row
        metric_label = METRIC_LABELS.get(metric, metric)
        cell = ws.cell(row=current_row, column=1, value=metric_label)
        cell.font = Font(bold=True, size=11, color="374151")
        fill = METRIC_FILL.get(metric, TOTAL_FILL)
        for c in range(1, total_col + 1):
            ws.cell(row=current_row, column=c).fill = fill
        current_row += 1

        metric_rows = [r for r in rows if r["metric"] == metric]
        for r in metric_rows:
            ws.cell(row=current_row, column=1, value=r["label"]).font = LABEL_FONT
            ws.cell(row=current_row, column=1).alignment = LEFT

            font = METRIC_FONT.get(metric, PCT_FONT)
            is_pct = metric in ("t2b", "b2b")

            for i, bc in enumerate(banner_columns):
                col_data = r.get("columns", {}).get(bc.letter, {})
                val = col_data.get("value")
                sig = col_data.get("sig_letters", [])

                if val is None:
                    ws.cell(row=current_row, column=2 + i, value="-").font = COUNT_FONT
                else:
                    display = f"{val:.1f}%" if is_pct else f"{val:.2f}"
                    if sig:
                        display += " " + "".join(sig)
                        cell = ws.cell(row=current_row, column=2 + i, value=display)
                        cell.font = SIG_FONT
                    else:
                        cell = ws.cell(row=current_row, column=2 + i, value=display)
                        cell.font = font
                ws.cell(row=current_row, column=2 + i).alignment = CENTER

            # Total
            total_val = r.get("total")
            if total_val is not None:
                display = f"{total_val:.1f}%" if is_pct else f"{total_val:.2f}"
                cell = ws.cell(row=current_row, column=total_col, value=display)
                cell.font = font
                cell.fill = TOTAL_FILL
                cell.alignment = CENTER

            for c in range(1, total_col + 1):
                ws.cell(row=current_row, column=c).border = THIN_BORDER
            current_row += 1

        current_row += 1  # Gap between metric groups

    # Column widths
    ws.column_dimensions["A"].width = 40
    for i in range(n_cols + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16
    ws.freeze_panes = f"B{letter_row + 1}"


STUB_HEADER_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
STUB_HEADER_FONT = Font(bold=True, size=11, color="1F2937")
STUB_SEPARATOR = Border(top=Side(style="medium", color="374151"))


def _write_single_sheet(ws, result: TabulationResult, spec: TabulateSpec, data: Any):
    """Write all stubs stacked on a single sheet.

    Layout:
    Row 1: Title
    Row 2: Sig note
    Row 3: Banner group headers (if multi)
    Row 4: Column value labels
    Row 5: Column letters
    Row 6+: Stubs stacked, each with header + base + data + nets + means
    """
    banner_columns = result.banner_columns
    banners = spec.resolved_banners
    n_cols = len(banner_columns)
    last_col = 1 + n_cols
    col_labels = getattr(data.meta, "column_names_to_labels", {})

    # Row 1: Title
    title = spec.title or "Tabulation Report"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="1F2937")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    # Row 2: Sig note
    sig_note = f"Significance: {spec.significance_level:.0%} confidence"
    if spec.weight:
        sig_note += f" | Weighted by: {spec.weight}"
    ws.cell(row=2, column=1, value=sig_note).font = Font(italic=True, size=9, color="6B7280")

    # Row 3: Banner group headers
    has_multi = len(banners) > 1
    offset = 1 if has_multi else 0
    if has_multi:
        col_idx = 2
        for b_idx, bvar in enumerate(banners):
            group_cols = [bc for bc in banner_columns if bc.banner_index == b_idx]
            if group_cols:
                cell = ws.cell(row=3, column=col_idx, value=group_cols[0].banner_label)
                cell.font = Font(bold=True, size=10, color="FFFFFF")
                cell.fill = BANNER_GROUP_FILL
                cell.alignment = CENTER
                if len(group_cols) > 1:
                    ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx + len(group_cols) - 1)
                col_idx += len(group_cols)

    # Column value labels
    label_row = 3 + offset
    for i, bc in enumerate(banner_columns):
        is_total = bc.banner_var == "_total_"
        cell = ws.cell(row=label_row, column=2 + i, value=bc.value_label)
        cell.font = HEADER_FONT
        cell.fill = TOTAL_FILL if is_total else HEADER_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER

    # Letters
    letter_row = label_row + 1
    for i, bc in enumerate(banner_columns):
        cell = ws.cell(row=letter_row, column=2 + i, value=bc.letter or "")
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = LETTER_HEADER_FILL
        cell.alignment = CENTER

    current_row = letter_row + 1

    # ── Stack each stub ──
    for sheet_result in result.sheets:
        if sheet_result.status != "success":
            continue

        # Grid sheets get their own treatment
        if sheet_result.is_grid and sheet_result.grid_data:
            current_row = _write_grid_block_in_single_sheet(
                ws, sheet_result, spec, banner_columns, current_row, last_col,
            )
            continue

        all_ct = sheet_result.crosstab_data
        if not all_ct:
            continue

        # ── Stub header ──
        stub_label = f"{sheet_result.variable}"
        if sheet_result.label and sheet_result.label != sheet_result.variable:
            stub_label += f": {sheet_result.label}"
        if sheet_result.is_mrs:
            stub_label += " (MRS)"

        for c in range(1, last_col + 1):
            ws.cell(row=current_row, column=c).border = STUB_SEPARATOR
        ws.cell(row=current_row, column=1, value=stub_label).font = STUB_HEADER_FONT
        for c in range(1, last_col + 1):
            ws.cell(row=current_row, column=c).fill = STUB_HEADER_FILL
        current_row += 1

        # ── Base row ──
        has_dual_base = spec.weight is not None
        base_label = "Base (weighted)" if has_dual_base else "Base (N)"
        ws.cell(row=current_row, column=1, value=base_label).font = BASE_FONT
        for i, bc in enumerate(banner_columns):
            ct = all_ct.get(bc.banner_var, {})
            base = sum(
                rd.get(bc.value, {}).get("count", 0)
                for rd in ct.get("table", []) if isinstance(rd.get(bc.value), dict)
            )
            cell = ws.cell(row=current_row, column=2 + i, value=int(round(base)))
            cell.font = BASE_FONT
            cell.alignment = CENTER
        current_row += 1

        if has_dual_base:
            ws.cell(row=current_row, column=1, value="Base (unweighted)").font = Font(size=9, color="9CA3AF")
            for i, bc in enumerate(banner_columns):
                try:
                    bval = float(bc.value) if bc.value.replace('.', '', 1).replace('-', '', 1).isdigit() else bc.value
                except (ValueError, AttributeError):
                    bval = bc.value
                uw = int((data.df[bc.banner_var] == bval).sum())
                cell = ws.cell(row=current_row, column=2 + i, value=uw)
                cell.font = Font(size=9, color="9CA3AF")
                cell.alignment = CENTER
            current_row += 1

        # ── Data rows ──
        all_rows = []
        seen_rv = set()
        for bvar in banners:
            ct = all_ct.get(bvar, {})
            for rd in ct.get("table", []):
                rv = rd.get("row_value")
                if rv not in seen_rv:
                    seen_rv.add(rv)
                    all_rows.append({"row_value": rv, "row_label": rd.get("row_label", str(rv))})

        for row_info in all_rows:
            rv = row_info["row_value"]
            ws.cell(row=current_row, column=1, value=row_info["row_label"]).font = LABEL_FONT

            for i, bc in enumerate(banner_columns):
                ct = all_ct.get(bc.banner_var, {})
                cell_data = None
                for rd in ct.get("table", []):
                    if rd.get("row_value") == rv:
                        cell_data = rd.get(bc.value, {})
                        break
                if not isinstance(cell_data, dict):
                    ws.cell(row=current_row, column=2 + i, value="-").font = COUNT_FONT
                    ws.cell(row=current_row, column=2 + i).alignment = CENTER
                    continue

                pct = cell_data.get("percentage", 0)
                orig_sig = cell_data.get("significance_letters", [])

                # Remap sig letters
                ct_cl = ct.get("col_labels", {})
                rev_map = {v: k for k, v in ct_cl.items()}
                global_sig = []
                for ol in orig_sig:
                    ov = rev_map.get(ol)
                    if ov:
                        for gbc in banner_columns:
                            if gbc.banner_var == bc.banner_var and gbc.value == ov:
                                global_sig.append(gbc.letter)
                                break

                is_total_col = bc.banner_var == "_total_"
                pct_str = f"{pct:.1f}%"
                if global_sig and not is_total_col:
                    pct_str += " " + "".join(global_sig)
                    cell = ws.cell(row=current_row, column=2 + i, value=pct_str)
                    cell.font = SIG_FONT
                else:
                    cell = ws.cell(row=current_row, column=2 + i, value=pct_str)
                    cell.font = PCT_FONT
                cell.alignment = CENTER
                if is_total_col:
                    cell.fill = TOTAL_FILL

            for c in range(1, last_col + 1):
                ws.cell(row=current_row, column=c).border = THIN_BORDER
            current_row += 1

        # ── Nets ──
        var_nets = (spec.nets or {}).get(sheet_result.variable, {})
        for net_name, net_values in var_nets.items():
            ws.cell(row=current_row, column=1, value=f"  {net_name}").font = Font(bold=True, size=9, color="166534")
            for i, bc in enumerate(banner_columns):
                ct = all_ct.get(bc.banner_var, {})
                net_count = 0
                col_base = sum(rd.get(bc.value, {}).get("count", 0) for rd in ct.get("table", []) if isinstance(rd.get(bc.value), dict))
                for rd in ct.get("table", []):
                    rv = rd.get("row_value")
                    if rv in net_values or (isinstance(rv, float) and int(rv) in net_values):
                        cd = rd.get(bc.value, {})
                        if isinstance(cd, dict):
                            net_count += cd.get("count", 0)
                net_pct = round(net_count / col_base * 100, 1) if col_base > 0 else 0
                cell = ws.cell(row=current_row, column=2 + i, value=f"{net_pct:.1f}%")
                cell.font = Font(size=9, color="166534")
                cell.fill = NET_FILL
                cell.alignment = CENTER
            current_row += 1

        # ── Means ──
        if spec.include_means and not sheet_result.is_mrs:
            stub_var = sheet_result.variable
            if pd.api.types.is_numeric_dtype(data.df[stub_var]):
                alpha = 1 - spec.significance_level
                ws.cell(row=current_row, column=1, value="  Mean").font = Font(bold=True, size=9, color="9A3412")
                for i, bc in enumerate(banner_columns):
                    try:
                        bval = float(bc.value) if bc.value.replace('.', '', 1).replace('-', '', 1).isdigit() else bc.value
                    except (ValueError, AttributeError):
                        bval = bc.value
                    mask = (data.df[bc.banner_var] == bval) & data.df[stub_var].notna()
                    vals = data.df[stub_var][mask]
                    if len(vals) < 1:
                        ws.cell(row=current_row, column=2 + i, value="-").font = COUNT_FONT
                        ws.cell(row=current_row, column=2 + i).alignment = CENTER
                        continue
                    m = float(vals.mean())
                    cell = ws.cell(row=current_row, column=2 + i, value=f"{m:.2f}")
                    cell.font = Font(size=9, color="9A3412")
                    cell.fill = MEAN_FILL
                    cell.alignment = CENTER
                current_row += 1

        current_row += 1  # Gap between stubs

    # Column widths
    ws.column_dimensions["A"].width = 45
    for i in range(n_cols):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16
    ws.freeze_panes = f"B{letter_row + 1}"


def _write_grid_block_in_single_sheet(ws, sheet_result, spec, banner_columns, start_row, last_col):
    """Write a grid summary block within the single-sheet layout."""
    grid = sheet_result.grid_data
    if not grid:
        return start_row

    rows = grid.get("rows", [])
    METRIC_LABELS = {"t2b": "T2B %", "b2b": "B2B %", "mean": "Mean", "median": "Median"}

    # Header
    for c in range(1, last_col + 1):
        ws.cell(row=start_row, column=c).border = STUB_SEPARATOR
        ws.cell(row=start_row, column=c).fill = STUB_HEADER_FILL
    label = sheet_result.label or sheet_result.variable
    ws.cell(row=start_row, column=1, value=f"{label} (Grid Summary)").font = STUB_HEADER_FONT
    start_row += 1

    metrics_seen = list(dict.fromkeys(r["metric"] for r in rows))
    for metric in metrics_seen:
        metric_rows = [r for r in rows if r["metric"] == metric]
        for r in metric_rows:
            ws.cell(row=start_row, column=1, value=f"  {r['label']} ({METRIC_LABELS.get(metric, metric)})").font = LABEL_FONT
            is_pct = metric in ("t2b", "b2b")
            for i, bc in enumerate(banner_columns):
                col_data = r.get("columns", {}).get(bc.letter, {})
                val = col_data.get("value")
                sig = col_data.get("sig_letters", [])
                if val is None:
                    ws.cell(row=start_row, column=2 + i, value="-").font = COUNT_FONT
                else:
                    display = f"{val:.1f}%" if is_pct else f"{val:.2f}"
                    if sig:
                        display += " " + "".join(sig)
                        ws.cell(row=start_row, column=2 + i, value=display).font = SIG_FONT
                    else:
                        ws.cell(row=start_row, column=2 + i, value=display).font = PCT_FONT
                ws.cell(row=start_row, column=2 + i).alignment = CENTER
            for c in range(1, last_col + 1):
                ws.cell(row=start_row, column=c).border = THIN_BORDER
            start_row += 1

    return start_row + 1
